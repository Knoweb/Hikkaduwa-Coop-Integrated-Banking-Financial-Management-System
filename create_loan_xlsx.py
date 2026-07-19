import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Loan Upload Template"

# Columns layout matching form flow
fields = [
    'loan_type_code', 'account_number', 'membership_number', 'member_nic',
    'requested_amount', 'term_months',
    'applied_date', 'disbursement_date',
    'status', 'loan_purpose', 'primary_job', 'annual_income_primary',
    
    # Guarantor 1 Details
    'guarantor1_name', 'guarantor1_nic', 'guarantor1_phone', 'guarantor1_address',
    'guarantor1_job', 'guarantor1_income',
    
    # Guarantor 2 Details
    'guarantor2_name', 'guarantor2_nic', 'guarantor2_phone', 'guarantor2_address',
    'guarantor2_job', 'guarantor2_income',
    
    # Defaults and System Generated
    'tenant_id', 'branch_id',
    'loan_id', 'created_at'
]

header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
header_font = Font(bold=True)

# Formatting colors
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # DO NOT FILL
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Leave Blank (Conditional)
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Example data cell

validations = {
    'status': DataValidation(type="list", formula1='"ACTIVE,PENDING,DISBURSED,CLOSED,REJECTED"', allow_blank=True),
    'loan_type_code': DataValidation(type="list", formula1='"පාරිභෝගික ණය,කෙටි ණය,සේවක ණය,ආපදා ණය,FD ණය (ස්ථාවර තැන්පතු ණය),අර්ත සාදක ණය,දුරකථන ණය,මහා සභා ණය,MPCS ණය (විවිධ සේවා සමූපාකාර සමිති ණය),කොටස් මත ණය,අත්තිකාරම් ණය,කල්පසු ණය,සමිතියේ කල්පසු ණය"', allow_blank=True)
}

for dv in validations.values():
    ws.add_data_validation(dv)

# Configure headers
for col_num, header in enumerate(fields, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = max(len(header) + 4, 12)
    
    if header in validations:
        validations[header].add(f'{col_letter}2:{col_letter}1000')
        
    # Format code/ID fields as Text to prevent scientific notation auto-conversion
    if header in ['account_number', 'membership_number', 'member_nic', 'guarantor1_nic', 'guarantor1_phone', 'guarantor2_nic', 'guarantor2_phone']:
        for row in range(2, 1001):
            ws.cell(row=row, column=col_num).number_format = '@'

# --- EXAMPLE ROWS ---

# Row 2: Disbursed Active Consumer Loan
data_row_2 = {
    'loan_type_code': 'පාරිභෝගික ණය', 'account_number': 'LN-HKW-2026-4728', 'membership_number': 'M-1-1001', 'member_nic': '199012345678',
    'requested_amount': 200000.00, 'term_months': 36,
    'applied_date': '2026-06-17', 'disbursement_date': '2026-06-22',
    'status': 'ACTIVE', 'loan_purpose': 'For buying a laptop', 'primary_job': 'Intern', 'annual_income_primary': 150000.00,
    'guarantor1_name': 'Amila Kumarasri', 'guarantor1_nic': '861453597V', 'guarantor1_phone': '0777442800', 'guarantor1_address': 'Baddegama', 'guarantor1_job': 'Manager', 'guarantor1_income': 500000.00,
    'guarantor2_name': 'Nihal Perera', 'guarantor2_nic': '6561320255V', 'guarantor2_phone': '0777267542', 'guarantor2_address': 'Kandy', 'guarantor2_job': 'Hotel', 'guarantor2_income': 600000.00,
    'tenant_id': 1, 'branch_id': 1,
    'loan_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL'
}

# Row 3: Pending Emergency Loan (No guarantor details needed, short term emergency)
data_row_3 = {
    'loan_type_code': 'ආපදා ණය', 'account_number': 'LN-HKW-2026-0002', 'membership_number': 'M-1-1002', 'member_nic': '198522334455',
    'requested_amount': 50000.00, 'term_months': 10,
    'applied_date': 'Leave Blank (Defaults to Today)', 'disbursement_date': 'Leave Blank (Pending)',
    'status': 'PENDING', 'loan_purpose': 'Disaster recovery', 'primary_job': 'Farmer', 'annual_income_primary': 80000.00,
    'guarantor1_name': 'Leave Blank (Optional)', 'guarantor1_nic': 'Leave Blank (Optional)', 'guarantor1_phone': 'Leave Blank (Optional)', 'guarantor1_address': 'Leave Blank (Optional)', 'guarantor1_job': 'Leave Blank (Optional)', 'guarantor1_income': 'Leave Blank (Optional)',
    'guarantor2_name': 'Leave Blank (Optional)', 'guarantor2_nic': 'Leave Blank (Optional)', 'guarantor2_phone': 'Leave Blank (Optional)', 'guarantor2_address': 'Leave Blank (Optional)', 'guarantor2_job': 'Leave Blank (Optional)', 'guarantor2_income': 'Leave Blank (Optional)',
    'tenant_id': 1, 'branch_id': 1,
    'loan_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL'
}

# Row 4: Disbursed Short Term Loan (One guarantor only)
data_row_4 = {
    'loan_type_code': 'කෙටි ණය', 'account_number': 'LN-HKW-2026-0003', 'membership_number': 'M-1-1003', 'member_nic': '197511223344',
    'requested_amount': 100000.00, 'term_months': 12,
    'applied_date': '2026-07-01', 'disbursement_date': '2026-07-02',
    'status': 'ACTIVE', 'loan_purpose': 'Business working capital', 'primary_job': 'Shop owner', 'annual_income_primary': 300000.00,
    'guarantor1_name': 'Sunil Shantha', 'guarantor1_nic': '723456789V', 'guarantor1_phone': '0711122334', 'guarantor1_address': 'Galle', 'guarantor1_job': 'Driver', 'guarantor1_income': 200000.00,
    'guarantor2_name': 'Leave Blank (Single)', 'guarantor2_nic': 'Leave Blank (Single)', 'guarantor2_phone': 'Leave Blank (Single)', 'guarantor2_address': 'Leave Blank (Single)', 'guarantor2_job': 'Leave Blank (Single)', 'guarantor2_income': 'Leave Blank (Single)',
    'tenant_id': 1, 'branch_id': 1,
    'loan_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL'
}

# Helper to write rows
def write_example_row(row_idx, data_dict):
    for col_idx, header in enumerate(fields, 1):
        val = data_dict.get(header)
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        
        # Color formatting
        if str(val).startswith("Leave Blank"):
            cell.fill = yellow_fill
            cell.font = Font(color="7F6000", italic=True)
            cell.alignment = Alignment(horizontal="center")
        elif val == "DO NOT FILL":
            cell.fill = red_fill
            cell.font = Font(color="C00000", bold=True)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.fill = green_fill
            cell.font = Font(color="375623")

write_example_row(2, data_row_2)
write_example_row(3, data_row_3)
write_example_row(4, data_row_4)

# --- SECOND SHEET: REFERENCE LEGEND ---
ws_ref = wb.create_sheet(title="Loan Types Reference")

ref_headers = ["Loan Type (ණය වර්ගය)", "Default Interest Rate (පොළී අනුපාතය)", "Max Term Months (උපරිම කාලය)", "Max Amount (උපරිම මුදල)"]
for col_num, header in enumerate(ref_headers, 1):
    cell = ws_ref.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    ws_ref.column_dimensions[get_column_letter(col_num)].width = 35

loan_types_data = [
    ("පාරිභෝගික ණය", "12.00%", "60", "Rs. 500,000.00"),
    ("කෙටි ණය", "14.00%", "12", "Rs. 100,000.00"),
    ("සේවක ණය", "10.00%", "36", "Rs. 300,000.00"),
    ("ආපදා ණය", "8.00%", "10", "Rs. 50,000.00"),
    ("FD ණය (ස්ථාවර තැන්පතු ණය)", "11.00%", "60", "Rs. 1,000,000.00"),
    ("අර්ත සාදක ණය", "9.00%", "60", "Rs. 500,000.00"),
    ("දුරකථන ණය", "15.00%", "24", "Rs. 100,000.00"),
    ("මහා සභා ණය", "12.00%", "36", "Rs. 200,000.00"),
    ("MPCS ණය (විවිධ සේවා සමූපාකාර සමිති ණය)", "10.00%", "60", "Rs. 2,000,000.00"),
    ("කොටස් මත ණය", "12.00%", "12", "Rs. 50,000.00"),
    ("අත්තිකාරම් ණය", "0.00%", "6", "Rs. 25,000.00"),
    ("කල්පසු ණය", "24.00%", "0", "නැත"),
    ("සමිතියේ කල්පසු ණය", "24.00%", "0", "නැත")
]

for row_idx, data in enumerate(loan_types_data, 2):
    for col_idx, val in enumerate(data, 1):
        ws_ref.cell(row=row_idx, column=col_idx, value=val)

# Save as v4
wb.save('loan_upload_template_final_v4.xlsx')
print("Loan Excel template V4 created successfully.")
