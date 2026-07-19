import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fixed Deposit Template"

# Reordered fields to match the UI form logic (dropdowns first)
fields = [
    'fd_type_code', 'account_mode', 'interest_payout_method', 'maturity_instruction',
    'fd_number', 'opened_date', 'principal_amount', 'linked_savings_account_number',
    'membership_number', 'member_nic', 'joint_membership_number_2', 'joint_membership_number_3',
    'occupation1', 'occupation2', 'occupation3',
    'receipt_number', 'has_submitted_tax_form',
    
    # Defaults and System Generated fields at the end
    'tenant_id', 'branch_id', 'status',
    'fd_id', 'created_at'
]

header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
header_font = Font(bold=True)

# Formatting colors
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # DO NOT FILL
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Leave Blank (Conditional)
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Example data cell

validations = {
    'account_mode': DataValidation(type="list", formula1='"single,joint"', allow_blank=True),
    'interest_payout_method': DataValidation(type="list", formula1='"AT_MATURITY,MONTHLY"', allow_blank=True),
    'maturity_instruction': DataValidation(type="list", formula1='"REINVEST_PRINCIPAL_AND_INTEREST,REINVEST_PRINCIPAL_PAY_INTEREST,CLOSE_ACCOUNT"', allow_blank=True),
    'has_submitted_tax_form': DataValidation(type="list", formula1='"true,false"', allow_blank=True),
    'status': DataValidation(type="list", formula1='"ACTIVE,CLOSED"', allow_blank=True),
    'fd_type_code': DataValidation(type="list", formula1='"FD_NRM_3M,FD_NRM_6M,FD_NRM_1Y,FD_NRM_24M,FD_NRM_60M,FD_SNR_3M,FD_SNR_6M,FD_SNR_1Y,FD_SNR_24M,FD_SNR_60M,FD_CHD_3M,FD_CHD_6M,FD_CHD_12M,FD_CHD_24M,FD_CHD_5Y,FD_JAY_3M,FD_JAY_6M,FD_JAY_1Y,FD_JAY_24M,FD_JAY_60M,FD_SIY_3M,FD_SIY_1Y,FD_SIY_60M"', allow_blank=True)
}

for dv in validations.values():
    ws.add_data_validation(dv)

# Configure headers
for col_num, header in enumerate(fields, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = max(len(header) + 4, 12)
    
    if header in validations:
        validations[header].add(f'{col_letter}2:{col_letter}1000')
        
    if header in ['fd_number', 'membership_number', 'member_nic', 'joint_membership_number_2', 'joint_membership_number_3', 'linked_savings_account_number']:
        for row in range(2, 1001):
            ws.cell(row=row, column=col_num).number_format = '@'

# --- EXAMPLE ROWS ---

# Row 2: Single Adult (New FD Account)
data_row_2 = {
    'fd_type_code': 'FD_NRM_1Y', 'account_mode': 'single', 'interest_payout_method': 'AT_MATURITY', 'maturity_instruction': 'REINVEST_PRINCIPAL_AND_INTEREST',
    'fd_number': 'FD-00001', 'opened_date': 'Leave Blank (Defaults to Today)', 'principal_amount': 50000.00, 'linked_savings_account_number': 'Leave Blank (No Payout)',
    'membership_number': 'M-1-1001', 'member_nic': '199012345678', 
    'joint_membership_number_2': 'Leave Blank (Single)', 'joint_membership_number_3': 'Leave Blank (Single)',
    'occupation1': 'Teacher', 'occupation2': 'Leave Blank (Single)', 'occupation3': 'Leave Blank (Single)',
    'receipt_number': 'REC-99901', 'has_submitted_tax_form': 'false',
    'tenant_id': 1, 'branch_id': 1, 'status': 'ACTIVE',
    'fd_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL'
}

# Row 3: Joint Senior Citizen (Old Migrated FD Account)
data_row_3 = {
    'fd_type_code': 'FD_SNR_1Y', 'account_mode': 'joint', 'interest_payout_method': 'MONTHLY', 'maturity_instruction': 'REINVEST_PRINCIPAL_PAY_INTEREST',
    'fd_number': 'FD-00002', 'opened_date': '2024-05-10', 'principal_amount': 250000.00, 'linked_savings_account_number': 'ACC-0001',
    'membership_number': 'M-1-1002', 'member_nic': '195598765432', 
    'joint_membership_number_2': 'M-1-1003', 'joint_membership_number_3': 'Leave Blank',
    'occupation1': 'Retired', 'occupation2': 'Housewife', 'occupation3': 'Leave Blank',
    'receipt_number': 'REC-99902', 'has_submitted_tax_form': 'true',
    'tenant_id': 1, 'branch_id': 1, 'status': 'ACTIVE',
    'fd_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL'
}

# Row 4: Single Child (New FD Account)
data_row_4 = {
    'fd_type_code': 'FD_CHD_5Y', 'account_mode': 'single', 'interest_payout_method': 'AT_MATURITY', 'maturity_instruction': 'CLOSE_ACCOUNT',
    'fd_number': 'FD-00003', 'opened_date': 'Leave Blank (Defaults to Today)', 'principal_amount': 10000.00, 'linked_savings_account_number': 'ACC-0003',
    'membership_number': 'M-1-1004', 'member_nic': '198011223344', 
    'joint_membership_number_2': 'Leave Blank (Single)', 'joint_membership_number_3': 'Leave Blank (Single)',
    'occupation1': 'Student', 'occupation2': 'Leave Blank (Single)', 'occupation3': 'Leave Blank (Single)',
    'receipt_number': 'REC-99903', 'has_submitted_tax_form': 'false',
    'tenant_id': 1, 'branch_id': 1, 'status': 'ACTIVE',
    'fd_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL'
}

# Helper to write rows
def write_example_row(row_idx, data_dict):
    for col_idx, header in enumerate(fields, 1):
        val = data_dict.get(header)
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        
        # Color formatting
        if str(val).startswith("Leave Blank"):
            cell.fill = yellow_fill
            cell.font = Font(color="7F6000", italic=True)
            cell.alignment = Alignment(horizontal="center")
        elif val == "DO NOT FILL":
            cell.fill = red_fill
            cell.font = Font(color="C00000", bold=True)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.fill = green_fill
            cell.font = Font(color="375623")

write_example_row(2, data_row_2)
write_example_row(3, data_row_3)
write_example_row(4, data_row_4)

# --- SECOND SHEET: REFERENCE LEGEND ---
ws_ref = wb.create_sheet(title="FD Types Reference")

ref_headers = ["FD Type Code", "Meaning (සිංහල තේරුම)", "Interest Maturity (කල්පිරෙන විට)", "Interest Monthly (මාසිකව)"]
for col_num, header in enumerate(ref_headers, 1):
    cell = ws_ref.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    ws_ref.column_dimensions[get_column_letter(col_num)].width = 30

fd_types_data = [
    ("FD_NRM_3M", "සාමාන්‍ය ස්ථාවර තැන්පතු - මාස 3", "10.00%", "8.00%"),
    ("FD_NRM_6M", "සාමාන්‍ය ස්ථාවර තැන්පතු - මාස 6", "17.00%", "15.00%"),
    ("FD_NRM_1Y", "සාමාන්‍ය ස්ථාවර තැන්පතු - අවුරුදු 1", "12.00%", "10.00%"),
    ("FD_NRM_24M", "සාමාන්‍ය ස්ථාවර තැන්පතු - අවුරුදු 2", "15.00%", "13.00%"),
    ("FD_NRM_60M", "සාමාන්‍ය ස්ථාවර තැන්පතු - අවුරුදු 5", "15.00%", "13.00%"),
    
    ("FD_SNR_3M", "ජ්‍යෙෂ්ඨ පුරවැසි තැන්පතු - මාස 3", "0.00%", "0.00%"),
    ("FD_SNR_6M", "ජ්‍යෙෂ්ඨ පුරවැසි තැන්පතු - මාස 6", "18.00%", "16.00%"),
    ("FD_SNR_1Y", "ජ්‍යෙෂ්ඨ පුරවැසි තැන්පතු - අවුරුදු 1", "13.00%", "11.00%"),
    ("FD_SNR_24M", "ජ්‍යෙෂ්ඨ පුරවැසි තැන්පතු - අවුරුදු 2", "0.00%", "0.00%"),
    ("FD_SNR_60M", "ජ්‍යෙෂ්ඨ පුරවැසි තැන්පතු - අවුරුදු 5", "0.00%", "0.00%"),
    
    ("FD_CHD_3M", "ළමා ස්ථාවර තැන්පතු - මාස 3", "0.00%", "0.00%"),
    ("FD_CHD_6M", "ළමා ස්ථාවර තැන්පතු - මාස 6", "0.00%", "0.00%"),
    ("FD_CHD_12M", "ළමා ස්ථාවර තැන්පතු - මාස 12", "0.00%", "0.00%"),
    ("FD_CHD_24M", "ළමා ස්ථාවර තැන්පතු - අවුරුදු 2", "0.00%", "0.00%"),
    ("FD_CHD_5Y", "ළමා ස්ථාවර තැන්පතු - අවුරුදු 5", "14.00%", "12.00%"),
    
    ("FD_JAY_3M", "ජය ස්ථාවර තැන්පතු - මාස 3", "10.00%", "8.00%"),
    ("FD_JAY_6M", "ජය ස්ථාවර තැන්පතු - මාස 6", "17.00%", "15.00%"),
    ("FD_JAY_1Y", "ජය ස්ථාවර තැන්පතු - අවුරුදු 1", "12.00%", "10.00%"),
    ("FD_JAY_24M", "ජය ස්ථාවර තැන්පතු - අවුරුදු 2", "15.00%", "13.00%"),
    ("FD_JAY_60M", "ජය ස්ථාවර තැන්පතු - අවුරුදු 5", "15.00%", "13.00%"),
    
    ("FD_SIY_3M", "සියවස් ස්ථාවර තැන්පතු - මාස 3", "10.00%", "8.00%"),
    ("FD_SIY_1Y", "සියවස් ස්ථාවර තැන්පතු - අවුරුදු 1", "12.00%", "10.00%"),
    ("FD_SIY_60M", "සියවස් ස්ථාවර තැන්පතු - අවුරුදු 5", "15.00%", "13.00%")
]

for row_idx, data in enumerate(fd_types_data, 2):
    for col_idx, val in enumerate(data, 1):
        ws_ref.cell(row=row_idx, column=col_idx, value=val)

wb.save('fd_upload_template_final_v3.xlsx')
print("Fixed Deposit Excel file with Reference sheet created successfully.")
