import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Members Template"

fields = [
    'tenant_id', 'member_id', 'membership_number', 'nic', 
    'name_with_initials', 'full_name', 'date_of_birth', 'gender', 
    'marital_status', 'address', 'contact_number', 'is_member', 
    'registered_branch_id', 'status', 'created_at', 'age_category'
]

header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
header_font = Font(bold=True)

validations = {
    'gender': DataValidation(type="list", formula1='"MALE,FEMALE"', allow_blank=True),
    'marital_status': DataValidation(type="list", formula1='"MARRIED,UNMARRIED"', allow_blank=True),
    'is_member': DataValidation(type="list", formula1='"true,false"', allow_blank=True),
    'status': DataValidation(type="list", formula1='"ACTIVE,INACTIVE"', allow_blank=True),
    'age_category': DataValidation(type="list", formula1='"ADULT,MINOR"', allow_blank=True)
}

for dv in validations.values():
    ws.add_data_validation(dv)

for col_num, header in enumerate(fields, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    
    col_letter = get_column_letter(col_num)
    
    if header in ['address', 'full_name']:
        ws.column_dimensions[col_letter].width = 35
    elif header in ['member_id', 'created_at', 'name_with_initials']:
        ws.column_dimensions[col_letter].width = 25
    else:
        ws.column_dimensions[col_letter].width = 20
        
    if header in validations:
        validations[header].add(f'{col_letter}2:{col_letter}1000')
        
    # Set text format for numbers to avoid scientific notation
    if header in ['membership_number', 'nic', 'contact_number']:
        for row in range(2, 1001):
            ws.cell(row=row, column=col_num).number_format = '@'

wb.save('member_registration_template_final.xlsx')
print("Final Excel file with Text formatting created successfully.")
