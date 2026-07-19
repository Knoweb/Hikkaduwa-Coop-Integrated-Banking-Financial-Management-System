import openpyxl
import psycopg2
from psycopg2 import sql

DB_CONFIG = {
    'dbname': 'hmcs_db',
    'user': 'hmcs_app',
    'password': 'hmcs_secure_pass_2026',
    'host': 'localhost',
    'port': '5432'
}

def get_member_uuid(cur, membership_no, nic=None):
    if membership_no and str(membership_no).lower() != 'no' and 'no (if single)' not in str(membership_no).lower():
        cur.execute("SELECT member_id FROM member_service.members WHERE membership_number = %s;", (membership_no,))
        res = cur.fetchone()
        if res:
            return res[0]
            
    if nic and str(nic).lower() != 'no' and 'no (if single)' not in str(nic).lower():
        cur.execute("SELECT member_id FROM member_service.members WHERE nic = %s;", (str(nic),))
        res = cur.fetchone()
        if res:
            return res[0]
    return None

def get_interest_rate(cur, account_type):
    if not account_type:
        return 0.04
    cur.execute("SELECT interest_rate FROM account_service.savings_account_type WHERE code = %s LIMIT 1;", (account_type,))
    res = cur.fetchone()
    if res:
        return float(res[0])
    return 0.04

def import_savings():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    
    wb = openpyxl.load_workbook('savings_upload_template_final_v5.xlsx')
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = {headers[col_idx-1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers)+1)}
        
        # Check if row is empty
        if not row_data.get('account_number'):
            continue
            
        # Skip example/info rows
        if str(row_data.get('account_id')).strip() == 'DO NOT FILL' or row_data.get('account_number') == 'No':
            continue
            
        # Resolve UUIDs from membership numbers / NIC
        member_id = get_member_uuid(cur, row_data.get('membership_number'), row_data.get('member_nic'))
        if not member_id:
            print(f"Skipping row {row_idx}: Member {row_data.get('membership_number')} / {row_data.get('member_nic')} not found in database.")
            continue
            
        member_id_2 = get_member_uuid(cur, row_data.get('joint_membership_number_2'))
        member_id_3 = get_member_uuid(cur, row_data.get('joint_membership_number_3'))
        
        # Prepare fields for insertion
        is_member = True if str(row_data.get('is_member')).lower() == 'true' else False
        initial_deposit = float(row_data.get('initial_deposit') or 0.0)
        balance = float(row_data.get('balance') or initial_deposit)
        branch_id = int(row_data.get('branch_id') or 1)
        tenant_id = int(row_data.get('tenant_id') or 1)
        
        child_name = row_data.get('child_name')
        if child_name and 'no (if adult)' in child_name.lower():
            child_name = None
            
        child_birth_certificate = row_data.get('child_birth_certificate')
        if child_birth_certificate and 'no (if adult)' in child_birth_certificate.lower():
            child_birth_certificate = None
            
        child_date_of_birth = row_data.get('child_date_of_birth')
        if child_date_of_birth and 'no (if adult)' in str(child_date_of_birth).lower():
            child_date_of_birth = None

        opened_date = row_data.get('opened_date')
        if not opened_date or 'do not fill' in str(opened_date).lower() or str(opened_date).strip() == '':
            opened_date = None

        interest_rate = get_interest_rate(cur, row_data.get('account_type'))

        cur.execute("""
            INSERT INTO account_service.savings_accounts (
                account_id, tenant_id, account_number, account_type, balance, branch_id, 
                child_birth_certificate, child_date_of_birth, child_name, member_id, 
                member_id_2, member_id_3, mode_of_operation, status, witness_address, 
                witness_name, occupation1, occupation2, occupation3, initial_deposit, 
                account_mode, annual_interest_rate, opened_date, created_at
            ) VALUES (
                gen_random_uuid(), %s, %s, %s, %s, %s, 
                %s, %s, %s, %s, 
                %s, %s, %s, %s, %s, 
                %s, %s, %s, %s, %s, 
                %s, %s, COALESCE(%s, CURRENT_DATE), CURRENT_TIMESTAMP
            ) ON CONFLICT (account_number) DO NOTHING;
        """, (
            tenant_id, row_data.get('account_number'), row_data.get('account_type'), balance, branch_id,
            child_birth_certificate, child_date_of_birth, child_name, member_id,
            member_id_2, member_id_3, row_data.get('mode_of_operation'), row_data.get('status') or 'ACTIVE',
            row_data.get('witness_address'), row_data.get('witness_name'),
            row_data.get('occupation1'), row_data.get('occupation2'), row_data.get('occupation3'),
            initial_deposit, row_data.get('account_mode'), 
            interest_rate,
            opened_date
        ))
        
    conn.commit()
    cur.close()
    conn.close()
    print("Savings accounts imported successfully.")

if __name__ == '__main__':
    import_savings()
