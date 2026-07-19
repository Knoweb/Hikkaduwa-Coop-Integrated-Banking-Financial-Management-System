import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Savings Template"

# Reordered fields to match the UI form logic (dropdowns first)
fields = [
    'is_member', 'account_type', 'account_mode', 'mode_of_operation',
    'account_number', 'opened_date', 'initial_deposit', 'balance',
    'membership_number', 'member_nic', 'joint_membership_number_2', 'joint_membership_number_3',
    'occupation1', 'occupation2', 'occupation3',
    'witness_name', 'witness_address',
    'child_name', 'child_birth_certificate', 'child_date_of_birth',
    
    # Defaults and System Generated fields at the end
    'tenant_id', 'branch_id', 'status',
    'account_id', 'created_at'
]

header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
header_font = Font(bold=True)

# Formatting colors
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # DO NOT FILL
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Leave Blank (Conditional)
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Example data cell

validations = {
    'is_member': DataValidation(type="list", formula1='"true,false"', allow_blank=True),
    'account_mode': DataValidation(type="list", formula1='"single,joint"', allow_blank=True),
    'mode_of_operation': DataValidation(type="list", formula1='"self,any_one,all_joint,joint_three"', allow_blank=True),
    'status': DataValidation(type="list", formula1='"ACTIVE,INACTIVE,CLOSED"', allow_blank=True),
    'account_type': DataValidation(type="list", formula1='"samanaya,janasetha,dhana_yojana,vandana,arunalu,ranthilina"', allow_blank=True)
}

for dv in validations.values():
    ws.add_data_validation(dv)

# Configure headers
for col_num, header in enumerate(fields, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    
    col_letter = get_column_letter(col_num)
    
    # Column Widths
    if header in ['witness_address', 'witness_name', 'child_name']:
        ws.column_dimensions[col_letter].width = 35
    elif header in ['account_id', 'member_id', 'member_id_2', 'member_id_3', 'membership_number', 'joint_membership_number_2', 'joint_membership_number_3']:
        ws.column_dimensions[col_letter].width = 30
    elif header in ['account_number', 'occupation1', 'occupation2', 'occupation3', 'member_nic']:
        ws.column_dimensions[col_letter].width = 25
    else:
        ws.column_dimensions[col_letter].width = 20
        
    if header in validations:
        validations[header].add(f'{col_letter}2:{col_letter}1000')
        
    if header in ['account_number', 'child_birth_certificate', 'membership_number', 'member_nic', 'joint_membership_number_2', 'joint_membership_number_3']:
        for row in range(2, 1001):
            ws.cell(row=row, column=col_num).number_format = '@'

# --- EXAMPLE ROWS ---

# Row 2: Single Adult (New Account)
data_row_2 = {
    'is_member': 'true', 'account_type': 'samanaya', 'account_mode': 'single', 'mode_of_operation': 'self',
    'account_number': 'ACC-0001', 'initial_deposit': 5000.00, 'balance': 'Leave Blank (New Account)',
    'membership_number': 'M-1-1001', 'member_nic': '199012345678', 
    'joint_membership_number_2': 'Leave Blank (Single)', 'joint_membership_number_3': 'Leave Blank (Single)',
    'occupation1': 'Teacher', 'occupation2': 'Leave Blank (Single)', 'occupation3': 'Leave Blank (Single)',
    'witness_name': 'W. Sunil Perera', 'witness_address': 'No. 12, Galle Road, Hikkaduwa',
    'child_name': 'Leave Blank (Adult)', 'child_birth_certificate': 'Leave Blank (Adult)', 'child_date_of_birth': 'Leave Blank (Adult)',
    'tenant_id': 1, 'branch_id': 1, 'status': 'ACTIVE',
    'account_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL', 'opened_date': 'Leave Blank (Defaults to Today)'
}

# Row 3: Joint Adult (Old / Migrated Account)
data_row_3 = {
    'is_member': 'true', 'account_type': 'janasetha', 'account_mode': 'joint', 'mode_of_operation': 'any_one',
    'account_number': 'ACC-0002', 'initial_deposit': 0.00, 'balance': 75000.00,
    'membership_number': 'M-1-1002', 'member_nic': '198598765432', 
    'joint_membership_number_2': 'M-1-1003', 'joint_membership_number_3': 'Leave Blank',
    'occupation1': 'Businessman', 'occupation2': 'Housewife', 'occupation3': 'Leave Blank',
    'witness_name': 'W. Sunil Perera', 'witness_address': 'No. 12, Galle Road, Hikkaduwa',
    'child_name': 'Leave Blank (Adult)', 'child_birth_certificate': 'Leave Blank (Adult)', 'child_date_of_birth': 'Leave Blank (Adult)',
    'tenant_id': 1, 'branch_id': 1, 'status': 'ACTIVE',
    'account_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL', 'opened_date': '2021-03-25'
}

# Row 4: Single Child (New Account)
data_row_4 = {
    'is_member': 'true', 'account_type': 'arunalu', 'account_mode': 'single', 'mode_of_operation': 'self',
    'account_number': 'ACC-0003', 'initial_deposit': 1000.00, 'balance': 'Leave Blank (New Account)',
    'membership_number': 'M-1-1004', 'member_nic': '198011223344', 
    'joint_membership_number_2': 'Leave Blank (Single)', 'joint_membership_number_3': 'Leave Blank (Single)',
    'occupation1': 'Student', 'occupation2': 'Leave Blank (Single)', 'occupation3': 'Leave Blank (Single)',
    'witness_name': 'W. Sunil Perera', 'witness_address': 'No. 12, Galle Road, Hikkaduwa',
    'child_name': 'Baby Perera', 'child_birth_certificate': 'BC-998877', 'child_date_of_birth': '2015-05-15',
    'tenant_id': 1, 'branch_id': 1, 'status': 'ACTIVE',
    'account_id': 'DO NOT FILL', 'created_at': 'DO NOT FILL', 'opened_date': 'Leave Blank (Defaults to Today)'
}

# Helper to write rows
def write_example_row(row_idx, data_dict):
    for col_idx, header in enumerate(fields, 1):
        val = data_dict.get(header)
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        
        # Color formatting
        if str(val).startswith("Leave Blank"):
            cell.fill = yellow_fill
            cell.font = Font(color="7F6000", italic=True)
            cell.alignment = Alignment(horizontal="center")
        elif val == "DO NOT FILL":
            cell.fill = red_fill
            cell.font = Font(color="C00000", bold=True)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.fill = green_fill
            cell.font = Font(color="375623")

write_example_row(2, data_row_2)
write_example_row(3, data_row_3)
write_example_row(4, data_row_4)

wb.save('savings_upload_template_final_v5.xlsx')
print("Savings Excel file with clean examples created successfully.")
