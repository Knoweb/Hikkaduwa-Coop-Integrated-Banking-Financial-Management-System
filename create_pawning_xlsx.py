import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pawn Tickets Upload"

# Columns layout matching schema requirements
fields = [
    'ticket_number', 'membership_number', 'member_nic',
    'gross_weight_grams', 'net_weight_grams', 'purity_karat',
    'advance_amount', 'issue_date',
    'status', 'article_description', 'tenant_id', 'branch_id',
    
    # Defaults and System Generated
    'ticket_id', 'valuer_id'
]

header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
header_font = Font(bold=True)

# Formatting colors
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # DO NOT FILL
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Leave Blank (Conditional)
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Example data cell

validations = {
    'status': DataValidation(type="list", formula1='"ACTIVE,REDEEMED,OVERDUE,AUCTIONED"', allow_blank=True),
    'purity_karat': DataValidation(type="list", formula1='"24,22,21,20,18"', allow_blank=True)
}

for dv in validations.values():
    ws.add_data_validation(dv)

# Configure headers
for col_num, header in enumerate(fields, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = max(len(header) + 4, 12)
    
    if header in validations:
        validations[header].add(f'{col_letter}2:{col_letter}1000')
        
    # Format text fields to prevent Excel conversions
    if header in ['ticket_number', 'membership_number', 'member_nic']:
        for row in range(2, 1001):
            ws.cell(row=row, column=col_num).number_format = '@'

# --- EXAMPLE ROWS ---

# Row 2: Disbursed Active Pawn Chain
data_row_2 = {
    'ticket_number': '698594', 'membership_number': 'M-1-1001', 'member_nic': '199012345678',
    'gross_weight_grams': 15.00, 'net_weight_grams': 10.00, 'purity_karat': 22,
    'advance_amount': 250000.00, 'issue_date': '2026-06-24',
    'status': 'ACTIVE', 'article_description': 'Gold Chain', 'tenant_id': 1, 'branch_id': 1,
    'ticket_id': 'DO NOT FILL', 'valuer_id': 'DO NOT FILL'
}

# Row 3: Redeemed Gold Ring
data_row_3 = {
    'ticket_number': '698595', 'membership_number': 'M-1-1002', 'member_nic': '198522334455',
    'gross_weight_grams': 5.00, 'net_weight_grams': 3.00, 'purity_karat': 22,
    'advance_amount': 80000.00, 'issue_date': '2026-01-15',
    'status': 'REDEEMED', 'article_description': 'Gold Ring', 'tenant_id': 1, 'branch_id': 1,
    'ticket_id': 'DO NOT FILL', 'valuer_id': 'DO NOT FILL'
}

# Row 4: Overdue Gold Bangle
data_row_4 = {
    'ticket_number': '698596', 'membership_number': 'M-1-1003', 'member_nic': '197511223344',
    'gross_weight_grams': 20.00, 'net_weight_grams': 14.00, 'purity_karat': 18,
    'advance_amount': 200000.00, 'issue_date': '2025-05-10',
    'status': 'OVERDUE', 'article_description': 'Gold Bangle', 'tenant_id': 1, 'branch_id': 1,
    'ticket_id': 'DO NOT FILL', 'valuer_id': 'DO NOT FILL'
}

# Helper to write rows
def write_example_row(row_idx, data_dict):
    for col_idx, header in enumerate(fields, 1):
        val = data_dict.get(header)
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        
        # Color formatting
        if str(val).startswith("Leave Blank"):
            cell.fill = yellow_fill
            cell.font = Font(color="7F6000", italic=True)
            cell.alignment = Alignment(horizontal="center")
        elif val == "DO NOT FILL":
            cell.fill = red_fill
            cell.font = Font(color="C00000", bold=True)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.fill = green_fill
            cell.font = Font(color="375623")

write_example_row(2, data_row_2)
write_example_row(3, data_row_3)
write_example_row(4, data_row_4)

# Save workbook
wb.save('pawning_upload_template_final_v3.xlsx')
print("Pawning Excel template V3 created successfully.")
