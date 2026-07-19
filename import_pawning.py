import openpyxl
import psycopg2
from psycopg2 import sql
import datetime
import dateutil.relativedelta

DB_CONFIG = {
    'dbname': 'hmcs_db',
    'user': 'hmcs_app',
    'password': 'hmcs_secure_pass_2026',
    'host': 'localhost',
    'port': '5432'
}

def get_member_uuid(cur, membership_no, nic=None):
    if membership_no and str(membership_no).lower() != 'no' and 'leave blank' not in str(membership_no).lower():
        cur.execute("SELECT member_id FROM member_service.members WHERE membership_number = %s;", (membership_no,))
        res = cur.fetchone()
        if res:
            return res[0]
            
    if nic and str(nic).lower() != 'no' and 'leave blank' not in str(nic).lower():
        cur.execute("SELECT member_id FROM member_service.members WHERE nic = %s;", (str(nic),))
        res = cur.fetchone()
        if res:
            return res[0]
    return None

def import_pawn_tickets():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    
    wb = openpyxl.load_workbook('pawning_upload_template_final_v3.xlsx')
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = {headers[col_idx-1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers)+1)}
        
        # Check if row is empty
        if not row_data.get('ticket_number'):
            continue
            
        # Skip example/info rows
        if str(row_data.get('ticket_id')).strip() == 'DO NOT FILL' or row_data.get('ticket_number') == 'No':
            continue
            
        ticket_number = str(row_data.get('ticket_number')).strip()
        
        # Skip duplicate checking
        cur.execute("SELECT ticket_id FROM pawning_service.pawn_tickets WHERE ticket_number = %s;", (ticket_number,))
        if cur.fetchone():
            print(f"Row {row_idx}: Skipping. Pawn ticket '{ticket_number}' already exists in database.")
            continue
            
        # 1. Resolve Member UUID
        member_id = get_member_uuid(cur, row_data.get('membership_number'), row_data.get('member_nic'))
        if not member_id:
            print(f"Error Row {row_idx}: Member profile not found for {row_data.get('membership_number')} / {row_data.get('member_nic')}")
            continue
            
        # 2. Parse weights and values
        gross_weight = float(row_data.get('gross_weight_grams') or 0.0)
        net_weight = float(row_data.get('net_weight_grams') or 0.0)
        purity_karat = int(row_data.get('purity_karat') or 22)
        advance_amount = float(row_data.get('advance_amount') or 0.0)
        assessed_value = advance_amount # Setting assessed_value equal to advance_amount
        
        # Interest rate defaults to 13.00%
        interest_rate = 13.00
        
        # Dates
        issue_date = row_data.get('issue_date')
        if not issue_date or 'leave blank' in str(issue_date).lower() or str(issue_date).strip() == '':
            issue_date = datetime.date.today()
        elif isinstance(issue_date, str):
            issue_date = datetime.datetime.strptime(issue_date.strip(), '%Y-%m-%d').date()
        elif isinstance(issue_date, datetime.datetime):
            issue_date = issue_date.date()
            
        # Automatically calculate expiry_date as exactly 1 year after issue_date
        expiry_date = issue_date + dateutil.relativedelta.relativedelta(years=1)
            
        status = row_data.get('status') or 'ACTIVE'
        article_description = row_data.get('article_description') or ''
        branch_id = int(row_data.get('branch_id') or 1)
        valuer_id = '00000000-0000-0000-0000-000000000000'
        
        # Insert Pawn Ticket
        cur.execute("""
            INSERT INTO pawning_service.pawn_tickets (
                ticket_id, ticket_number, member_id, gross_weight_grams, net_weight_grams, purity_karat,
                assessed_value, advance_amount, interest_rate, branch_id, valuer_id,
                issue_date, expiry_date, status, article_description
            ) VALUES (
                gen_random_uuid(), %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s
            );
        """, (
            ticket_number, member_id, gross_weight, net_weight, purity_karat,
            assessed_value, advance_amount, interest_rate, branch_id, valuer_id,
            issue_date, expiry_date, status, article_description
        ))
        
    conn.commit()
    cur.close()
    conn.close()
    print("Pawn tickets imported successfully.")

if __name__ == '__main__':
    import_pawn_tickets()
