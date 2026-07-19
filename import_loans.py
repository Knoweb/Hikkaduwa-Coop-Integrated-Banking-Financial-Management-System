import openpyxl
import psycopg2
from psycopg2 import sql
import datetime
import json

DB_CONFIG = {
    'dbname': 'hmcs_db',
    'user': 'hmcs_app',
    'password': 'hmcs_secure_pass_2026',
    'host': 'localhost',
    'port': '5432'
}

def get_member_uuid_and_details(cur, membership_no, nic=None):
    if membership_no and str(membership_no).lower() != 'no' and 'leave blank' not in str(membership_no).lower():
        cur.execute("SELECT member_id, nic, mobile_number, full_name, membership_number, date_of_birth, gender FROM member_service.members WHERE membership_number = %s;", (membership_no,))
        res = cur.fetchone()
        if res:
            return res
            
    if nic and str(nic).lower() != 'no' and 'leave blank' not in str(nic).lower():
        cur.execute("SELECT member_id, nic, mobile_number, full_name, membership_number, date_of_birth, gender FROM member_service.members WHERE nic = %s;", (str(nic),))
        res = cur.fetchone()
        if res:
            return res
    return None

def get_loan_type_info(cur, Sinhala_name):
    cur.execute("SELECT loan_type_id, interest_rate, max_term_months, description FROM loan_service.loan_types WHERE name = %s;", (Sinhala_name,))
    res = cur.fetchone()
    if res:
        return {
            'loan_type_id': res[0],
            'interest_rate': float(res[1]) if res[1] else 12.00,
            'max_term_months': int(res[2]) if res[2] else 36,
            'description': res[3]
        }
    return None

def insert_guarantor(cur, loan_id, g_number, g_name, g_nic, g_phone, g_address, g_job, g_income):
    if not g_name or str(g_name).strip() == '' or 'leave blank' in str(g_name).lower():
        return
    income = float(g_income) if g_income and 'leave blank' not in str(g_income).lower() else None
    
    cur.execute("""
        INSERT INTO loan_service.loan_guarantors (
            loan_id, guarantor_id, full_name, nic, phone, address, job, annual_income_primary, guarantor_number
        ) VALUES (
            %s, gen_random_uuid(), %s, %s, %s, %s, %s, %s, %s
        );
    """, (loan_id, g_name, g_nic, g_phone, g_address, g_job, income, g_number))

def import_loans():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    
    wb = openpyxl.load_workbook('loan_upload_template_final_v4.xlsx')
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = {headers[col_idx-1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers)+1)}
        
        # Check if row is empty
        if not row_data.get('account_number'):
            continue
            
        # Skip example/info rows
        if str(row_data.get('loan_id')).strip() == 'DO NOT FILL' or row_data.get('account_number') == 'No':
            continue
            
        account_number = str(row_data.get('account_number')).strip()
        
        # Skip duplicate checking
        cur.execute("SELECT loan_id FROM loan_service.loans WHERE account_number = %s;", (account_number,))
        if cur.fetchone():
            print(f"Row {row_idx}: Skipping. Loan account '{account_number}' already exists in database.")
            continue
            
        # 1. Resolve Member UUID
        m_details = get_member_uuid_and_details(cur, row_data.get('membership_number'), row_data.get('member_nic'))
        if not m_details:
            print(f"Error Row {row_idx}: Member profile not found for {row_data.get('membership_number')} / {row_data.get('member_nic')}")
            continue
            
        member_id, applicant_nic, applicant_phone, applicant_name, applicant_member_no, dob, gender = m_details
        
        # 2. Resolve Loan Type Details
        loan_type_code = row_data.get('loan_type_code')
        type_info = get_loan_type_info(cur, loan_type_code)
        if not type_info:
            print(f"Error Row {row_idx}: Loan Type '{loan_type_code}' not found in database!")
            continue
            
        loan_type_id = type_info['loan_type_id']
        default_rate = type_info['interest_rate']
        
        interest_rate = default_rate
            
        # Amounts
        requested_amount = float(row_data.get('requested_amount') or 0.0)
        
        disbursed_amount = requested_amount
            
        term_months = int(row_data.get('term_months') or type_info['max_term_months'])
        
        # Dates
        applied_date = row_data.get('applied_date')
        if not applied_date or 'leave blank' in str(applied_date).lower() or str(applied_date).strip() == '':
            applied_date = datetime.date.today()
        elif isinstance(applied_date, str):
            applied_date = datetime.datetime.strptime(applied_date.strip(), '%Y-%m-%d').date()
        elif isinstance(applied_date, datetime.datetime):
            applied_date = applied_date.date()
            
        disb_date_val = row_data.get('disbursement_date')
        if not disb_date_val or 'leave blank' in str(disb_date_val).lower() or str(disb_date_val).strip() == '':
            disbursement_date = None
        elif isinstance(disb_date_val, str):
            disbursement_date = datetime.datetime.strptime(disb_date_val.strip(), '%Y-%m-%d')
        elif isinstance(disb_date_val, datetime.date):
            disbursement_date = datetime.datetime.combine(disb_date_val, datetime.time.min)
        elif isinstance(disb_date_val, datetime.datetime):
            disbursement_date = disb_date_val
            

            
        status = row_data.get('status') or 'ACTIVE'
        if status in ['ACTIVE', 'DISBURSED']:
            current_stage = 'DISBURSED'
        else:
            current_stage = 'STAGE_1_FIELD_VERIFICATION'
            
        tenant_id = int(row_data.get('tenant_id') or 1)
        branch_id = int(row_data.get('branch_id') or 1)
        
        # Default disbursed_by based on branch_id
        if branch_id == 1:
            disbursed_by = 'mgr_hkw'
        elif branch_id == 10:
            disbursed_by = 'mgr_sandarawala'
        else:
            disbursed_by = 'system_migration'
        
        # 3. Construct application_data JSONB
        guarantor1_data = {}
        if row_data.get('guarantor1_name') and 'leave blank' not in str(row_data.get('guarantor1_name')).lower():
            guarantor1_data = {
                "name": row_data.get('guarantor1_name'),
                "nic": row_data.get('guarantor1_nic'),
                "phone": row_data.get('guarantor1_phone'),
                "address": row_data.get('guarantor1_address'),
                "job": row_data.get('guarantor1_job'),
                "incomePrimary": str(row_data.get('guarantor1_income') or '')
            }
            
        guarantor2_data = {}
        if row_data.get('guarantor2_name') and 'leave blank' not in str(row_data.get('guarantor2_name')).lower():
            guarantor2_data = {
                "name": row_data.get('guarantor2_name'),
                "nic": row_data.get('guarantor2_nic'),
                "phone": row_data.get('guarantor2_phone'),
                "address": row_data.get('guarantor2_address'),
                "job": row_data.get('guarantor2_job'),
                "incomePrimary": str(row_data.get('guarantor2_income') or '')
            }
            
        app_json = {
            "dob": dob.isoformat() if dob else "",
            "nic": applicant_nic or "",
            "phone": applicant_phone or "",
            "gender": gender or "",
            "memberNo": applicant_member_no or "",
            "applicantName": applicant_name or "",
            "loanPurpose": row_data.get('loan_purpose') or "",
            "repaymentPeriodMonths": str(term_months),
            "primaryJob": row_data.get('primary_job') or "",
            "annualIncomePrimary": str(row_data.get('annual_income_primary') or ''),
            "guarantor1": guarantor1_data,
            "guarantor2": guarantor2_data
        }
        
        # Insert Loan and return UUID
        cur.execute("""
            INSERT INTO loan_service.loans (
                loan_id, member_id, loan_type, requested_amount, approved_amount, interest_rate, term_months,
                branch_id, current_stage, status, applied_date, application_data, 
                account_number, disbursement_date, disbursed_amount, disbursed_by,
                created_at, updated_at, loan_type_id
            ) VALUES (
                gen_random_uuid(), %s, %s, %s, NULL, %s, %s,
                %s, %s, %s, %s, %s::jsonb,
                %s, %s, %s, %s,
                CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, %s
            ) RETURNING loan_id;
        """, (
            member_id, type_info['description'], requested_amount, interest_rate, term_months,
            branch_id, current_stage, status, applied_date, json.dumps(app_json),
            account_number, disbursement_date, disbursed_amount, disbursed_by,
            loan_type_id
        ))
        loan_id = cur.fetchone()[0]
        
        # 4. Insert Guarantors
        insert_guarantor(cur, loan_id, 1, 
                         row_data.get('guarantor1_name'), row_data.get('guarantor1_nic'), 
                         row_data.get('guarantor1_phone'), row_data.get('guarantor1_address'), 
                         row_data.get('guarantor1_job'), row_data.get('guarantor1_income'))
                         
        insert_guarantor(cur, loan_id, 2, 
                         row_data.get('guarantor2_name'), row_data.get('guarantor2_nic'), 
                         row_data.get('guarantor2_phone'), row_data.get('guarantor2_address'), 
                         row_data.get('guarantor2_job'), row_data.get('guarantor2_income'))
                         
    conn.commit()
    cur.close()
    conn.close()
    print("Loan accounts and guarantors imported successfully.")

if __name__ == '__main__':
    import_loans()
