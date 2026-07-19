import openpyxl
import psycopg2
from psycopg2 import sql
import datetime

DB_CONFIG = {
    'dbname': 'hmcs_db',
    'user': 'hmcs_app',
    'password': 'hmcs_secure_pass_2026',
    'host': 'localhost',
    'port': '5432'
}

def get_member_uuid(cur, membership_no, nic=None):
    if membership_no and str(membership_no).lower() != 'no' and 'leave blank' not in str(membership_no).lower():
        cur.execute("SELECT member_id FROM member_service.members WHERE membership_number = %s;", (membership_no,))
        res = cur.fetchone()
        if res:
            return res[0]
            
    if nic and str(nic).lower() != 'no' and 'leave blank' not in str(nic).lower():
        cur.execute("SELECT member_id FROM member_service.members WHERE nic = %s;", (str(nic),))
        res = cur.fetchone()
        if res:
            return res[0]
    return None

def get_savings_uuid(cur, savings_acc_no):
    if savings_acc_no and str(savings_acc_no).lower() != 'no' and 'leave blank' not in str(savings_acc_no).lower():
        cur.execute("SELECT account_id FROM account_service.savings_accounts WHERE account_number = %s;", (savings_acc_no,))
        res = cur.fetchone()
        if res:
            return res[0]
    return None

def get_fd_type_info(cur, fd_type_code):
    cur.execute("SELECT id, interest_rate_maturity, interest_rate_monthly, term_months FROM account_service.fixed_deposit_types WHERE code = %s;", (fd_type_code,))
    res = cur.fetchone()
    if res:
        return {
            'type_id': res[0],
            'interest_rate_maturity': float(res[1]),
            'interest_rate_monthly': float(res[2]),
            'term_months': int(res[3])
        }
    return None

def add_months(sourcedate, months):
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, [31,
        29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28,
        31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month-1])
    return datetime.date(year, month, day)

def import_fd():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    
    wb = openpyxl.load_workbook('fd_upload_template_final_v3.xlsx')
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = {headers[col_idx-1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers)+1)}
        
        # Check if row is empty
        if not row_data.get('fd_number'):
            continue
            
        # Skip example/info rows
        if str(row_data.get('fd_id')).strip() == 'DO NOT FILL' or row_data.get('fd_number') == 'No':
            continue
            
        # 1. Resolve Member UUIDs
        member_id = get_member_uuid(cur, row_data.get('membership_number'), row_data.get('member_nic'))
        if not member_id:
            print(f"Error Row {row_idx}: Member not found for {row_data.get('membership_number')} / {row_data.get('member_nic')}")
            continue
            
        member_id_2 = get_member_uuid(cur, row_data.get('joint_membership_number_2'))
        member_id_3 = get_member_uuid(cur, row_data.get('joint_membership_number_3'))
        
        # 2. Resolve Linked Savings Account
        linked_savings_account_id = get_savings_uuid(cur, row_data.get('linked_savings_account_number'))
        
        # 3. Resolve FD Type details (Interest rate and Term)
        fd_type_code = row_data.get('fd_type_code')
        type_info = get_fd_type_info(cur, fd_type_code)
        if not type_info:
            print(f"Error Row {row_idx}: FD Type Code '{fd_type_code}' not found in database!")
            continue
            
        # Resolve interest rate depending on payout method
        payout_method = row_data.get('interest_payout_method') or 'AT_MATURITY'
        if payout_method == 'MONTHLY':
            interest_rate = type_info['interest_rate_monthly']
        else:
            interest_rate = type_info['interest_rate_maturity']
            
        term_months = type_info['term_months']
        type_id = type_info['type_id']
        
        # 4. Dates handling
        opened_date = row_data.get('opened_date')
        if not opened_date or 'leave blank' in str(opened_date).lower() or str(opened_date).strip() == '':
            opened_date = datetime.date.today()
        elif isinstance(opened_date, str):
            opened_date = datetime.datetime.strptime(opened_date.strip(), '%Y-%m-%d').date()
        elif isinstance(opened_date, datetime.datetime):
            opened_date = opened_date.date()
            
        # Calculate maturity date: opened_date + term_months
        maturity_date = add_months(opened_date, term_months)
        
        # 5. Numeric values
        principal_amount = float(row_data.get('principal_amount') or 0.0)
        
        has_submitted_tax_form = str(row_data.get('has_submitted_tax_form')).lower() == 'true'
        tenant_id = int(row_data.get('tenant_id') or 1)
        branch_id = int(row_data.get('branch_id') or 1)
        
        cur.execute("""
            INSERT INTO account_service.fixed_deposits (
                fd_id, principal_amount, interest_rate, term_months, maturity_date, status, 
                fd_number, interest_payout_method, linked_savings_account_id, maturity_instruction, 
                member_id, receipt_number, member_id_2, member_id_3, type_id, 
                is_officer_approved, last_interest_payout_date, accumulated_interest, 
                opened_date, has_submitted_tax_form, branch_id, created_at
            ) VALUES (
                gen_random_uuid(), %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                true, NULL, 0.00,
                %s, %s, %s, CURRENT_TIMESTAMP
            ) ON CONFLICT (fd_number) DO NOTHING;
        """, (
            principal_amount, interest_rate, term_months, maturity_date, row_data.get('status') or 'ACTIVE',
            row_data.get('fd_number'), payout_method, linked_savings_account_id, row_data.get('maturity_instruction'),
            member_id, row_data.get('receipt_number'), member_id_2, member_id_3, type_id,
            opened_date, has_submitted_tax_form, branch_id
        ))
        
    conn.commit()
    cur.close()
    conn.close()
    print("Fixed deposit accounts imported successfully.")

if __name__ == '__main__':
    import_fd()
